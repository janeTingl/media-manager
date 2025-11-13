"""Tests for import/export functionality."""

import json
from datetime import datetime
from pathlib import Path

import pytest
from openpyxl import load_workbook
from sqlmodel import Session, SQLModel, create_engine
from sqlmodel.pool import StaticPool

from media_manager.import_export_service import (
    ExportFormat,
    ExportOptions,
    ImportExportService,
    ImportOptions,
    MergeStrategy,
)
from media_manager.persistence.database import DatabaseService
from media_manager.persistence.models import (
    ExternalId,
    Library,
    MediaFile,
    MediaItem,
)
from media_manager.persistence.repositories import transactional_context


@pytest.fixture
def in_memory_db():
    """Create an in-memory SQLite database for testing."""
    engine = create_engine(
        "sqlite://",
        echo=False,
        connect_args={"check_same_thread": False},
        poolclass=StaticPool,
    )

    SQLModel.metadata.create_all(engine)

    db_service = DatabaseService("sqlite://", auto_migrate=False)
    db_service._engine = engine

    import media_manager.persistence.database as db_module
    db_module._database_service = db_service

    session = Session(engine)

    yield db_service, session

    session.close()
    engine.dispose()
    db_module._database_service = None


@pytest.fixture
def service():
    """Create import/export service."""
    return ImportExportService()


@pytest.fixture
def sample_library(in_memory_db):
    """Create a sample library."""
    _, session = in_memory_db
    library = Library(
        name="Test Library",
        path="/test/library",
        media_type="mixed",
        is_active=True,
    )
    session.add(library)
    session.commit()
    session.refresh(library)
    return library


@pytest.fixture
def sample_media_items(in_memory_db, sample_library):
    """Create sample media items."""
    _, session = in_memory_db
    items = []

    # Create a movie
    movie = MediaItem(
        library_id=sample_library.id,
        title="Test Movie",
        media_type="movie",
        year=2023,
        description="A test movie",
        genres="Action, Sci-Fi",
        runtime=120,
        rating=8.5,
        created_at=datetime.utcnow(),
        updated_at=datetime.utcnow(),
    )
    session.add(movie)
    session.flush()

    # Add file
    movie_file = MediaFile(
        media_item_id=movie.id,
        path="/test/library/test_movie.mkv",
        filename="test_movie.mkv",
        file_size=1024 * 1024 * 1024,
        container="mkv",
        video_codec="h264",
        audio_codec="aac",
        resolution="1920x1080",
    )
    session.add(movie_file)

    # Add external ID
    tmdb_id = ExternalId(
        media_item_id=movie.id,
        source="tmdb",
        external_id="12345",
    )
    session.add(tmdb_id)

    items.append(movie)

    # Create a TV episode
    episode = MediaItem(
        library_id=sample_library.id,
        title="Test Show",
        media_type="tv",
        year=2023,
        season=1,
        episode=1,
        description="A test episode",
        genres="Drama",
        runtime=45,
        rating=7.8,
        created_at=datetime.utcnow(),
        updated_at=datetime.utcnow(),
    )
    session.add(episode)
    session.flush()

    # Add file
    episode_file = MediaFile(
        media_item_id=episode.id,
        path="/test/library/test_show_s01e01.mkv",
        filename="test_show_s01e01.mkv",
        file_size=512 * 1024 * 1024,
    )
    session.add(episode_file)

    items.append(episode)

    session.commit()
    for item in items:
        session.refresh(item)

    return items


class TestExportJSON:
    """Tests for JSON export."""

    def test_export_json_basic(self, service, sample_media_items, tmp_path):
        """Test basic JSON export."""
        output_file = tmp_path / "export.json"

        options = ExportOptions(
            format=ExportFormat.JSON,
            include_files=True,
            include_external_ids=True,
            include_artwork=False,
            include_subtitles=False,
        )

        count = service.export_to_file(output_file, options)
        assert count == 2

        # Verify file contents
        with open(output_file, "r", encoding="utf-8") as f:
            data = json.load(f)

        assert "version" in data
        assert "exported_at" in data
        assert "items" in data
        assert len(data["items"]) == 2

        # Check movie data
        movie = next(item for item in data["items"] if item["media_type"] == "movie")
        assert movie["title"] == "Test Movie"
        assert movie["year"] == 2023
        assert movie["genres"] == "Action, Sci-Fi"
        assert movie["runtime"] == 120
        assert movie["rating"] == 8.5
        assert len(movie["files"]) == 1
        assert movie["files"][0]["filename"] == "test_movie.mkv"
        assert len(movie["external_ids"]) == 1
        assert movie["external_ids"][0]["source"] == "tmdb"

        # Check TV episode data
        episode = next(item for item in data["items"] if item["media_type"] == "tv")
        assert episode["title"] == "Test Show"
        assert episode["season"] == 1
        assert episode["episode"] == 1

    def test_export_json_filtered_by_library(self, service, sample_media_items, tmp_path):
        """Test JSON export filtered by library."""
        output_file = tmp_path / "export_filtered.json"

        options = ExportOptions(
            format=ExportFormat.JSON,
            library_ids=[sample_media_items[0].library_id],
        )

        count = service.export_to_file(output_file, options)
        assert count == 2

    def test_export_json_filtered_by_media_type(self, service, sample_media_items, tmp_path):
        """Test JSON export filtered by media type."""
        output_file = tmp_path / "export_movies.json"

        options = ExportOptions(
            format=ExportFormat.JSON,
            media_types=["movie"],
        )

        count = service.export_to_file(output_file, options)
        assert count == 1

        with open(output_file, "r", encoding="utf-8") as f:
            data = json.load(f)

        assert len(data["items"]) == 1
        assert data["items"][0]["media_type"] == "movie"


class TestExportExcel:
    """Tests for Excel export."""

    def test_export_excel_basic(self, service, sample_media_items, tmp_path):
        """Test basic Excel export."""
        output_file = tmp_path / "export.xlsx"

        options = ExportOptions(
            format=ExportFormat.EXCEL,
            include_files=True,
            include_external_ids=True,
            include_artwork=True,
            include_subtitles=True,
        )

        count = service.export_to_file(output_file, options)
        assert count == 2

        # Verify file contents
        wb = load_workbook(output_file)
        ws = wb.active
        assert ws is not None

        # Check headers
        headers = [cell.value for cell in ws[1]]
        assert "Title" in headers
        assert "Media Type" in headers
        assert "Year" in headers
        assert "File Paths" in headers
        assert "TMDB ID" in headers

        # Check data rows (2 items + 1 header = 3 rows)
        assert ws.max_row == 3

        # Check movie row
        movie_row = None
        for row in ws.iter_rows(min_row=2, max_row=3):
            if row[headers.index("Title")].value == "Test Movie":
                movie_row = row
                break

        assert movie_row is not None
        assert movie_row[headers.index("Media Type")].value == "movie"
        assert movie_row[headers.index("Year")].value == 2023
        assert movie_row[headers.index("Runtime")].value == 120
        assert movie_row[headers.index("Rating")].value == 8.5

        wb.close()

    def test_export_excel_column_width(self, service, sample_media_items, tmp_path):
        """Test Excel export auto-sizes columns."""
        output_file = tmp_path / "export_columns.xlsx"

        options = ExportOptions(format=ExportFormat.EXCEL)
        service.export_to_file(output_file, options)

        wb = load_workbook(output_file)
        ws = wb.active
        assert ws is not None

        # Check that columns have been sized
        for column in ws.column_dimensions.values():
            assert column.width > 0

        wb.close()


class TestImportJSON:
    """Tests for JSON import."""

    def test_import_json_basic(self, service, sample_library, tmp_path):
        """Test basic JSON import."""
        import_file = tmp_path / "import.json"

        # Create import data
        data = {
            "version": "1.0",
            "exported_at": datetime.utcnow().isoformat(),
            "items": [
                {
                    "title": "Imported Movie",
                    "media_type": "movie",
                    "year": 2024,
                    "description": "An imported movie",
                    "genres": "Action",
                    "runtime": 90,
                    "rating": 7.5,
                    "files": [
                        {
                            "path": "/test/imported_movie.mkv",
                            "filename": "imported_movie.mkv",
                            "file_size": 1024 * 1024,
                        }
                    ],
                    "external_ids": [
                        {"source": "tmdb", "external_id": "99999"}
                    ],
                }
            ],
        }

        with open(import_file, "w", encoding="utf-8") as f:
            json.dump(data, f)

        # Import
        options = ImportOptions(
            target_library_id=sample_library.id,
            validate_files=False,
        )

        result = service.import_from_file(import_file, options)

        assert result.total == 1
        assert result.imported == 1
        assert result.failed == 0

        # Verify imported item
        with transactional_context() as uow:
            from sqlmodel import select

            stmt = select(MediaItem).where(MediaItem.title == "Imported Movie")
            item = uow.session.exec(stmt).first()

            assert item is not None
            assert item.media_type == "movie"
            assert item.year == 2024
            assert item.runtime == 90
            assert item.rating == 7.5
            assert len(item.files) == 1
            assert len(item.external_ids) == 1

    def test_import_json_with_conflicts_skip(self, service, sample_media_items, tmp_path):
        """Test JSON import with conflicts using skip strategy."""
        import_file = tmp_path / "import_conflict.json"

        # Create import data with existing item
        data = {
            "version": "1.0",
            "items": [
                {
                    "id": sample_media_items[0].id,
                    "title": sample_media_items[0].title,
                    "media_type": "movie",
                    "year": 2023,
                    "library_id": sample_media_items[0].library_id,
                }
            ],
        }

        with open(import_file, "w", encoding="utf-8") as f:
            json.dump(data, f)

        # Import with skip strategy
        options = ImportOptions(
            merge_strategy=MergeStrategy.SKIP,
            validate_files=False,
        )

        result = service.import_from_file(import_file, options)

        assert result.total == 1
        assert result.skipped == 1
        assert result.imported == 0

    def test_import_json_with_conflicts_update(self, service, sample_media_items, tmp_path):
        """Test JSON import with conflicts using update strategy."""
        import_file = tmp_path / "import_update.json"

        # Create import data with updated information
        data = {
            "version": "1.0",
            "items": [
                {
                    "id": sample_media_items[0].id,
                    "title": sample_media_items[0].title,
                    "media_type": "movie",
                    "year": 2023,
                    "description": "Updated description",
                    "rating": 9.0,
                    "library_id": sample_media_items[0].library_id,
                }
            ],
        }

        with open(import_file, "w", encoding="utf-8") as f:
            json.dump(data, f)

        # Import with update strategy
        options = ImportOptions(
            merge_strategy=MergeStrategy.UPDATE,
            validate_files=False,
        )

        result = service.import_from_file(import_file, options)

        assert result.total == 1
        assert result.updated == 1
        assert result.imported == 0

        # Verify update
        with transactional_context() as uow:
            from sqlmodel import select

            stmt = select(MediaItem).where(MediaItem.id == sample_media_items[0].id)
            item = uow.session.exec(stmt).first()

            assert item is not None
            assert item.description == "Updated description"
            assert item.rating == 9.0


class TestImportExcel:
    """Tests for Excel import."""

    def test_import_excel_basic(self, service, sample_library, tmp_path):
        """Test basic Excel import."""
        import_file = tmp_path / "import.xlsx"

        # Create Excel file
        from openpyxl import Workbook

        wb = Workbook()
        ws = wb.active
        assert ws is not None

        # Write headers
        headers = ["Title", "Media Type", "Year", "Description", "Genres", "Runtime", "Rating"]
        ws.append(headers)

        # Write data
        ws.append(["Imported Movie", "movie", 2024, "Test", "Action", 90, 7.5])
        ws.append(["Imported Show", "tv", 2024, "Test", "Drama", 45, 8.0])

        wb.save(import_file)

        # Import
        options = ImportOptions(
            target_library_id=sample_library.id,
            validate_files=False,
        )

        result = service.import_from_file(import_file, options)

        assert result.total == 2
        assert result.imported == 2
        assert result.failed == 0

    def test_import_excel_with_column_mapping(self, service, sample_library, tmp_path):
        """Test Excel import with custom column mapping."""
        import_file = tmp_path / "import_mapped.xlsx"

        # Create Excel file with different column names
        from openpyxl import Workbook

        wb = Workbook()
        ws = wb.active
        assert ws is not None

        # Write headers
        headers = ["Name", "Type", "Released", "Summary"]
        ws.append(headers)

        # Write data
        ws.append(["Mapped Movie", "movie", 2024, "A mapped movie"])

        wb.save(import_file)

        # Import with column mapping
        column_mapping = {
            "Name": "title",
            "Type": "media_type",
            "Released": "year",
            "Summary": "description",
        }

        options = ImportOptions(
            target_library_id=sample_library.id,
            validate_files=False,
        )

        result = service.import_from_file(import_file, options, column_mapping)

        assert result.total == 1
        assert result.imported == 1

        # Verify imported item
        with transactional_context() as uow:
            from sqlmodel import select

            stmt = select(MediaItem).where(MediaItem.title == "Mapped Movie")
            item = uow.session.exec(stmt).first()

            assert item is not None
            assert item.description == "A mapped movie"


class TestRoundTrip:
    """Tests for round-trip import/export."""

    def test_roundtrip_json(self, service, sample_media_items, tmp_path):
        """Test exporting and re-importing JSON data."""
        export_file = tmp_path / "roundtrip.json"

        # Export
        export_options = ExportOptions(
            format=ExportFormat.JSON,
            include_files=True,
            include_external_ids=True,
        )

        exported = service.export_to_file(export_file, export_options)
        assert exported == 2

        # Delete original items and related data
        with transactional_context() as uow:
            from sqlmodel import select
            
            for item in sample_media_items:
                # Delete related data first
                for ext_id in uow.session.exec(select(ExternalId).where(ExternalId.media_item_id == item.id)).all():
                    uow.session.delete(ext_id)
                for file in uow.session.exec(select(MediaFile).where(MediaFile.media_item_id == item.id)).all():
                    uow.session.delete(file)
                
                # Now delete the item
                stmt = select(MediaItem).where(MediaItem.id == item.id)
                item_to_delete = uow.session.exec(stmt).first()
                if item_to_delete:
                    uow.session.delete(item_to_delete)

        # Re-import
        import_options = ImportOptions(
            target_library_id=sample_media_items[0].library_id,
            validate_files=False,
        )

        result = service.import_from_file(export_file, import_options)

        assert result.total == 2
        assert result.imported == 2

    def test_roundtrip_excel(self, service, sample_media_items, tmp_path):
        """Test exporting and re-importing Excel data."""
        export_file = tmp_path / "roundtrip.xlsx"

        # Export
        export_options = ExportOptions(
            format=ExportFormat.EXCEL,
            include_files=True,
            include_external_ids=True,
        )

        exported = service.export_to_file(export_file, export_options)
        assert exported == 2

        # Delete original items and related data
        with transactional_context() as uow:
            from sqlmodel import select
            
            for item in sample_media_items:
                # Delete related data first
                for ext_id in uow.session.exec(select(ExternalId).where(ExternalId.media_item_id == item.id)).all():
                    uow.session.delete(ext_id)
                for file in uow.session.exec(select(MediaFile).where(MediaFile.media_item_id == item.id)).all():
                    uow.session.delete(file)
                
                # Now delete the item
                stmt = select(MediaItem).where(MediaItem.id == item.id)
                item_to_delete = uow.session.exec(stmt).first()
                if item_to_delete:
                    uow.session.delete(item_to_delete)

        # Re-import
        import_options = ImportOptions(
            target_library_id=sample_media_items[0].library_id,
            validate_files=False,
        )

        result = service.import_from_file(export_file, import_options)

        assert result.total == 2
        assert result.imported == 2


class TestErrorHandling:
    """Tests for error handling."""

    def test_import_missing_required_field(self, service, sample_library, tmp_path):
        """Test import with missing required field."""
        import_file = tmp_path / "import_missing.json"

        # Create import data without title
        data = {
            "version": "1.0",
            "items": [
                {
                    "media_type": "movie",
                    "year": 2024,
                }
            ],
        }

        with open(import_file, "w", encoding="utf-8") as f:
            json.dump(data, f)

        options = ImportOptions(
            target_library_id=sample_library.id,
            validate_files=False,
        )

        result = service.import_from_file(import_file, options)

        assert result.total == 1
        assert result.failed == 1
        assert len(result.errors) == 1

    def test_import_invalid_file_format(self, service, sample_library, tmp_path):
        """Test import with invalid file format."""
        import_file = tmp_path / "import.txt"
        import_file.write_text("invalid content")

        options = ImportOptions(target_library_id=sample_library.id)

        with pytest.raises(ValueError, match="Unsupported import format"):
            service.import_from_file(import_file, options)

    def test_import_nonexistent_library(self, service, in_memory_db, tmp_path):
        """Test import with non-existent library."""
        import_file = tmp_path / "import.json"

        data = {
            "version": "1.0",
            "items": [
                {
                    "title": "Test Movie",
                    "media_type": "movie",
                    "year": 2024,
                }
            ],
        }

        with open(import_file, "w", encoding="utf-8") as f:
            json.dump(data, f)

        options = ImportOptions(
            target_library_id=99999,  # Non-existent
            validate_files=False,
        )

        result = service.import_from_file(import_file, options)

        assert result.total == 1
        assert result.failed == 1

    def test_preview_import_conflicts(self, service, sample_media_items, tmp_path):
        """Test preview import to detect conflicts."""
        import_file = tmp_path / "import_preview.json"

        # Create import data with existing item
        data = {
            "version": "1.0",
            "items": [
                {
                    "title": sample_media_items[0].title,
                    "media_type": "movie",
                    "year": sample_media_items[0].year,
                }
            ],
        }

        with open(import_file, "w", encoding="utf-8") as f:
            json.dump(data, f)

        # Preview import
        data_preview, conflicts = service.preview_import(import_file)

        assert len(data_preview) == 1
        assert len(conflicts) >= 1
        assert any(c.conflict_type == "duplicate_title" for c in conflicts)

    def test_export_unsupported_format(self, service, in_memory_db, tmp_path):
        """Test export with unsupported format."""
        output_file = tmp_path / "export.csv"

        options = ExportOptions(format="csv")  # type: ignore

        with pytest.raises(ValueError, match="Unsupported export format"):
            service.export_to_file(output_file, options)
