"""Service for importing and exporting media metadata."""

from __future__ import annotations

import json
from dataclasses import dataclass, field
from datetime import datetime
from enum import Enum
from pathlib import Path
from typing import Any, Callable, Iterable, Optional

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill
from sqlalchemy.orm import selectinload
from sqlmodel import select

from .logging import get_logger
from .persistence.models import (
    Artwork,
    ExternalId,
    HistoryEvent,
    Library,
    MediaFile,
    MediaItem,
    Subtitle,
)
from .persistence.repositories import transactional_context


class ExportFormat(str, Enum):
    """Export format options."""

    JSON = "json"
    EXCEL = "excel"


class MergeStrategy(str, Enum):
    """Strategy for handling conflicts during import."""

    SKIP = "skip"  # Skip items that already exist
    REPLACE = "replace"  # Replace existing items completely
    UPDATE = "update"  # Update existing items with new data


@dataclass
class ExportOptions:
    """Options for export operation."""

    format: ExportFormat
    library_ids: list[int] = field(default_factory=list)
    media_types: list[str] = field(default_factory=list)  # "movie", "tv"
    include_files: bool = True
    include_external_ids: bool = True
    include_artwork: bool = True
    include_subtitles: bool = True
    date_from: datetime | None = None
    date_to: datetime | None = None


@dataclass
class ImportOptions:
    """Options for import operation."""

    merge_strategy: MergeStrategy = MergeStrategy.SKIP
    target_library_id: int | None = None
    validate_files: bool = True
    update_existing: bool = True
    create_history: bool = True


@dataclass
class ImportConflict:
    """Represents a conflict during import."""

    row_index: int
    title: str
    conflict_type: str  # "duplicate_id", "duplicate_title", "missing_library"
    existing_item_id: int | None = None
    details: str = ""


@dataclass
class ImportResult:
    """Result of an import operation."""

    total: int = 0
    imported: int = 0
    updated: int = 0
    skipped: int = 0
    failed: int = 0
    conflicts: list[ImportConflict] = field(default_factory=list)
    errors: list[str] = field(default_factory=list)

    def to_message(self) -> str:
        """Return a human-readable message describing the result."""
        parts: list[str] = []
        parts.append(f"{self.total} items processed")
        if self.imported:
            parts.append(f"{self.imported} imported")
        if self.updated:
            parts.append(f"{self.updated} updated")
        if self.skipped:
            parts.append(f"{self.skipped} skipped")
        if self.failed:
            parts.append(f"{self.failed} failed")
        if self.conflicts:
            parts.append(f"{len(self.conflicts)} conflicts")
        return ", ".join(parts)


class ImportExportService:
    """Service for importing and exporting media metadata."""

    def __init__(self) -> None:
        self._logger = get_logger().get_logger(__name__)

    def export_to_file(
        self,
        file_path: Path,
        options: ExportOptions,
        progress_callback: Optional[Callable[[int, int, str], None]] = None,
    ) -> int:
        """Export media items to a file."""
        with transactional_context() as uow:
            session = uow.session

            data = self._fetch_export_data_in_session(session, options)

            if progress_callback:
                progress_callback(0, len(data), "Starting export...")

            if options.format == ExportFormat.JSON:
                exported = self._export_to_json(file_path, data, options)
            elif options.format == ExportFormat.EXCEL:
                exported = self._export_to_excel(file_path, data, options, progress_callback)
            else:
                raise ValueError(f"Unsupported export format: {options.format}")

            if progress_callback:
                progress_callback(len(data), len(data), f"Export complete: {exported} items")

            self._logger.info(f"Exported {exported} items to {file_path}")
            return exported

    def import_from_file(
        self,
        file_path: Path,
        options: ImportOptions,
        column_mapping: dict[str, str] | None = None,
        progress_callback: Optional[Callable[[int, int, str], None]] = None,
    ) -> ImportResult:
        """Import media items from a file."""
        file_ext = file_path.suffix.lower()

        if file_ext == ".json":
            data = self._import_from_json(file_path)
        elif file_ext in (".xlsx", ".xls"):
            data = self._import_from_excel(file_path, column_mapping)
        else:
            raise ValueError(f"Unsupported import format: {file_ext}")

        result = self._process_import_data(data, options, progress_callback)
        self._logger.info(f"Import complete: {result.to_message()}")
        return result

    def preview_import(
        self,
        file_path: Path,
        column_mapping: dict[str, str] | None = None,
    ) -> tuple[list[dict[str, Any]], list[ImportConflict]]:
        """Preview import data and detect conflicts."""
        file_ext = file_path.suffix.lower()

        if file_ext == ".json":
            data = self._import_from_json(file_path)
        elif file_ext in (".xlsx", ".xls"):
            data = self._import_from_excel(file_path, column_mapping)
        else:
            raise ValueError(f"Unsupported import format: {file_ext}")

        conflicts = self._detect_conflicts(data)
        return data, conflicts

    def get_excel_headers(self, file_path: Path) -> list[str]:
        """Get column headers from an Excel file."""
        wb = load_workbook(file_path, read_only=True)
        ws = wb.active
        if ws is None:
            return []

        headers = []
        for cell in ws[1]:
            if cell.value:
                headers.append(str(cell.value))
        wb.close()
        return headers

    def _fetch_export_data_in_session(self, session: Any, options: ExportOptions) -> list[MediaItem]:
        """Fetch media items for export based on options within an existing session."""
        stmt = select(MediaItem).options(
            selectinload(MediaItem.library),
            selectinload(MediaItem.files),
            selectinload(MediaItem.external_ids),
            selectinload(MediaItem.artworks),
            selectinload(MediaItem.subtitles),
            selectinload(MediaItem.tags),
        )

        # Apply filters
        if options.library_ids:
            stmt = stmt.where(MediaItem.library_id.in_(options.library_ids))

        if options.media_types:
            stmt = stmt.where(MediaItem.media_type.in_(options.media_types))

        if options.date_from:
            stmt = stmt.where(MediaItem.created_at >= options.date_from)

        if options.date_to:
            stmt = stmt.where(MediaItem.created_at <= options.date_to)

        items = session.exec(stmt).all()
        return list(items)

    def _export_to_json(
        self,
        file_path: Path,
        items: list[MediaItem],
        options: ExportOptions,
    ) -> int:
        """Export media items to JSON format."""
        export_data = {
            "version": "1.0",
            "exported_at": datetime.utcnow().isoformat(),
            "items": [],
        }

        for item in items:
            item_data = self._serialize_media_item(item, options)
            export_data["items"].append(item_data)

        with open(file_path, "w", encoding="utf-8") as f:
            json.dump(export_data, f, indent=2, ensure_ascii=False)

        return len(items)

    def _export_to_excel(
        self,
        file_path: Path,
        items: list[MediaItem],
        options: ExportOptions,
        progress_callback: Optional[Callable[[int, int, str], None]] = None,
    ) -> int:
        """Export media items to Excel format."""
        wb = Workbook()
        ws = wb.active
        if ws is None:
            raise ValueError("Failed to create worksheet")

        ws.title = "Media Items"

        # Define headers
        headers = [
            "ID",
            "Library",
            "Title",
            "Media Type",
            "Year",
            "Season",
            "Episode",
            "Description",
            "Genres",
            "Runtime",
            "Aired Date",
            "Rating",
            "Created At",
            "Updated At",
        ]

        if options.include_files:
            headers.extend(["File Paths", "File Count"])

        if options.include_external_ids:
            headers.extend(["TMDB ID", "IMDB ID", "TVDB ID"])

        if options.include_artwork:
            headers.extend(["Poster Path", "Fanart Path", "Artwork Count"])

        if options.include_subtitles:
            headers.extend(["Subtitle Paths", "Subtitle Count"])

        # Write headers with styling
        header_font = Font(bold=True)
        header_fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")

        for col, header in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill

        # Write data rows
        for idx, item in enumerate(items, start=2):
            if progress_callback:
                progress_callback(idx - 1, len(items), f"Exporting {item.title}")

            row_data = [
                item.id,
                item.library.name if item.library else "",
                item.title,
                item.media_type,
                item.year,
                item.season,
                item.episode,
                item.description,
                item.genres,
                item.runtime,
                item.aired_date,
                item.rating,
                item.created_at.isoformat() if item.created_at else None,
                item.updated_at.isoformat() if item.updated_at else None,
            ]

            if options.include_files:
                file_paths = [f.path for f in item.files]
                row_data.append("; ".join(file_paths))
                row_data.append(len(item.files))

            if options.include_external_ids:
                ext_ids = {ext.source: ext.external_id for ext in item.external_ids}
                row_data.append(ext_ids.get("tmdb", ""))
                row_data.append(ext_ids.get("imdb", ""))
                row_data.append(ext_ids.get("tvdb", ""))

            if options.include_artwork:
                poster_path = ""
                fanart_path = ""
                for art in item.artworks:
                    if art.artwork_type == "poster" and art.local_path:
                        poster_path = art.local_path
                    elif art.artwork_type == "fanart" and art.local_path:
                        fanart_path = art.local_path
                row_data.append(poster_path)
                row_data.append(fanart_path)
                row_data.append(len(item.artworks))

            if options.include_subtitles:
                subtitle_paths = [s.local_path for s in item.subtitles if s.local_path]
                row_data.append("; ".join(subtitle_paths))
                row_data.append(len(item.subtitles))

            for col, value in enumerate(row_data, start=1):
                ws.cell(row=idx, column=col, value=value)

        # Auto-size columns
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                except:  # noqa: E722
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width

        wb.save(file_path)
        return len(items)

    def _serialize_media_item(
        self,
        item: MediaItem,
        options: ExportOptions,
    ) -> dict[str, Any]:
        """Serialize a media item to a dictionary."""
        data: dict[str, Any] = {
            "id": item.id,
            "library_id": item.library_id,
            "library_name": item.library.name if item.library else None,
            "title": item.title,
            "media_type": item.media_type,
            "year": item.year,
            "season": item.season,
            "episode": item.episode,
            "description": item.description,
            "genres": item.genres,
            "runtime": item.runtime,
            "aired_date": item.aired_date,
            "rating": item.rating,
            "created_at": item.created_at.isoformat() if item.created_at else None,
            "updated_at": item.updated_at.isoformat() if item.updated_at else None,
        }

        if options.include_files:
            data["files"] = [
                {
                    "path": f.path,
                    "filename": f.filename,
                    "file_size": f.file_size,
                    "duration": f.duration,
                    "container": f.container,
                    "video_codec": f.video_codec,
                    "audio_codec": f.audio_codec,
                    "resolution": f.resolution,
                }
                for f in item.files
            ]

        if options.include_external_ids:
            data["external_ids"] = [
                {"source": ext.source, "external_id": ext.external_id}
                for ext in item.external_ids
            ]

        if options.include_artwork:
            data["artworks"] = [
                {
                    "artwork_type": art.artwork_type,
                    "url": art.url,
                    "local_path": art.local_path,
                    "size": art.size,
                    "download_status": art.download_status,
                }
                for art in item.artworks
            ]

        if options.include_subtitles:
            data["subtitles"] = [
                {
                    "language": sub.language,
                    "format": sub.format,
                    "local_path": sub.local_path,
                    "provider": sub.provider,
                    "download_status": sub.download_status,
                }
                for sub in item.subtitles
            ]

        return data

    def _import_from_json(self, file_path: Path) -> list[dict[str, Any]]:
        """Import data from JSON format."""
        with open(file_path, "r", encoding="utf-8") as f:
            data = json.load(f)

        if isinstance(data, dict) and "items" in data:
            return data["items"]
        elif isinstance(data, list):
            return data
        else:
            raise ValueError("Invalid JSON format")

    def _import_from_excel(
        self,
        file_path: Path,
        column_mapping: dict[str, str] | None = None,
    ) -> list[dict[str, Any]]:
        """Import data from Excel format."""
        wb = load_workbook(file_path, read_only=True)
        ws = wb.active
        if ws is None:
            raise ValueError("No active worksheet found")

        # Read headers
        headers = []
        for cell in ws[1]:
            if cell.value:
                headers.append(str(cell.value))

        # Apply column mapping if provided
        if column_mapping:
            headers = [column_mapping.get(h, h) for h in headers]

        # Read data rows
        data = []
        for row in ws.iter_rows(min_row=2, values_only=True):
            if not any(row):
                continue

            item_data: dict[str, Any] = {}
            for header, value in zip(headers, row):
                if value is not None:
                    item_data[header] = value
            data.append(item_data)

        wb.close()
        return data

    def _process_import_data(
        self,
        data: list[dict[str, Any]],
        options: ImportOptions,
        progress_callback: Optional[Callable[[int, int, str], None]] = None,
    ) -> ImportResult:
        """Process and import data into the database."""
        result = ImportResult(total=len(data))

        with transactional_context() as uow:
            session = uow.session

            for idx, item_data in enumerate(data, start=1):
                if progress_callback:
                    title = item_data.get("title", "Unknown")
                    progress_callback(idx, len(data), f"Importing {title}")

                try:
                    self._import_single_item(
                        session, item_data, options, result, idx
                    )
                except Exception as e:
                    result.failed += 1
                    result.errors.append(f"Row {idx}: {str(e)}")
                    self._logger.error(f"Failed to import item at row {idx}: {e}")

        return result

    def _import_single_item(
        self,
        session: Any,
        item_data: dict[str, Any],
        options: ImportOptions,
        result: ImportResult,
        row_index: int,
    ) -> None:
        """Import a single media item."""
        title = item_data.get("title") or item_data.get("Title")
        if not title:
            raise ValueError("Missing required field: title")

        # Check for existing item by ID
        existing_item = None
        item_id = item_data.get("id") or item_data.get("ID")
        if item_id:
            stmt = select(MediaItem).where(MediaItem.id == int(item_id))
            existing_item = session.exec(stmt).first()

        # If no ID match, check by title and year
        if not existing_item:
            year = item_data.get("year") or item_data.get("Year")
            media_type = item_data.get("media_type") or item_data.get("Media Type")
            stmt = select(MediaItem).where(
                MediaItem.title == title,
                MediaItem.year == year if year else True,
                MediaItem.media_type == media_type if media_type else True,
            )
            existing_item = session.exec(stmt).first()

        # Handle conflicts based on merge strategy
        if existing_item:
            if options.merge_strategy == MergeStrategy.SKIP:
                result.skipped += 1
                return
            elif options.merge_strategy == MergeStrategy.REPLACE:
                self._delete_existing_item(session, existing_item)
                existing_item = None
            elif options.merge_strategy == MergeStrategy.UPDATE:
                self._update_existing_item(session, existing_item, item_data, options)
                result.updated += 1
                
                # Create history event for update
                if options.create_history and existing_item and existing_item.id:
                    history_event = HistoryEvent(
                        media_item_id=existing_item.id,
                        event_type="imported_update",
                        event_data=json.dumps({"source": "import", "row": row_index}),
                        timestamp=datetime.utcnow(),
                    )
                    session.add(history_event)
                return

        # Create new item
        new_item = self._create_new_item(session, item_data, options)
        result.imported += 1

        # Create history event
        if options.create_history and new_item and new_item.id:
            history_event = HistoryEvent(
                media_item_id=new_item.id,
                event_type="imported",
                event_data=json.dumps({"source": "import", "row": row_index}),
                timestamp=datetime.utcnow(),
            )
            session.add(history_event)

    def _create_new_item(
        self,
        session: Any,
        item_data: dict[str, Any],
        options: ImportOptions,
    ) -> MediaItem:
        """Create a new media item from import data."""
        # Determine library
        library_id = options.target_library_id
        if not library_id:
            library_id = item_data.get("library_id") or item_data.get("Library ID")

        if not library_id:
            raise ValueError("No target library specified")

        # Verify library exists
        stmt = select(Library).where(Library.id == int(library_id))
        library = session.exec(stmt).first()
        if not library:
            raise ValueError(f"Library with ID {library_id} not found")

        # Create media item
        media_item = MediaItem(
            library_id=int(library_id),
            title=item_data.get("title") or item_data.get("Title", ""),
            media_type=item_data.get("media_type") or item_data.get("Media Type", "movie"),
            year=self._parse_int(item_data.get("year") or item_data.get("Year")),
            season=self._parse_int(item_data.get("season") or item_data.get("Season")),
            episode=self._parse_int(item_data.get("episode") or item_data.get("Episode")),
            description=item_data.get("description") or item_data.get("Description"),
            genres=item_data.get("genres") or item_data.get("Genres"),
            runtime=self._parse_int(item_data.get("runtime") or item_data.get("Runtime")),
            aired_date=item_data.get("aired_date") or item_data.get("Aired Date"),
            rating=self._parse_float(item_data.get("rating") or item_data.get("Rating")),
            created_at=datetime.utcnow(),
            updated_at=datetime.utcnow(),
        )
        session.add(media_item)
        session.flush()

        # Add files if present
        files_data = item_data.get("files", [])
        if isinstance(files_data, str):
            # Handle semicolon-separated paths from Excel
            file_paths = [p.strip() for p in files_data.split(";") if p.strip()]
            for file_path in file_paths:
                if options.validate_files and not Path(file_path).exists():
                    self._logger.warning(f"File not found: {file_path}")
                    continue
                media_file = MediaFile(
                    media_item_id=media_item.id,
                    path=file_path,
                    filename=Path(file_path).name,
                    file_size=0,
                )
                session.add(media_file)
        elif isinstance(files_data, list):
            for file_info in files_data:
                if isinstance(file_info, dict):
                    file_path = file_info.get("path", "")
                    if options.validate_files and not Path(file_path).exists():
                        self._logger.warning(f"File not found: {file_path}")
                        continue
                    media_file = MediaFile(
                        media_item_id=media_item.id,
                        path=file_path,
                        filename=file_info.get("filename", Path(file_path).name),
                        file_size=file_info.get("file_size", 0),
                        duration=file_info.get("duration"),
                        container=file_info.get("container"),
                        video_codec=file_info.get("video_codec"),
                        audio_codec=file_info.get("audio_codec"),
                        resolution=file_info.get("resolution"),
                    )
                    session.add(media_file)

        # Add external IDs if present
        external_ids_data = item_data.get("external_ids", [])
        if isinstance(external_ids_data, list):
            for ext_id in external_ids_data:
                if isinstance(ext_id, dict):
                    external_id = ExternalId(
                        media_item_id=media_item.id,
                        source=ext_id.get("source", ""),
                        external_id=ext_id.get("external_id", ""),
                    )
                    session.add(external_id)
        else:
            # Handle individual ID columns from Excel
            tmdb_id = item_data.get("tmdb_id") or item_data.get("TMDB ID")
            imdb_id = item_data.get("imdb_id") or item_data.get("IMDB ID")
            tvdb_id = item_data.get("tvdb_id") or item_data.get("TVDB ID")

            if tmdb_id:
                session.add(ExternalId(media_item_id=media_item.id, source="tmdb", external_id=str(tmdb_id)))
            if imdb_id:
                session.add(ExternalId(media_item_id=media_item.id, source="imdb", external_id=str(imdb_id)))
            if tvdb_id:
                session.add(ExternalId(media_item_id=media_item.id, source="tvdb", external_id=str(tvdb_id)))

        return media_item

    def _update_existing_item(
        self,
        session: Any,
        existing_item: MediaItem,
        item_data: dict[str, Any],
        options: ImportOptions,
    ) -> None:
        """Update an existing media item with import data."""
        # Update basic fields
        if "title" in item_data or "Title" in item_data:
            existing_item.title = item_data.get("title") or item_data.get("Title", existing_item.title)
        if "description" in item_data or "Description" in item_data:
            existing_item.description = item_data.get("description") or item_data.get("Description")
        if "genres" in item_data or "Genres" in item_data:
            existing_item.genres = item_data.get("genres") or item_data.get("Genres")
        if "runtime" in item_data or "Runtime" in item_data:
            existing_item.runtime = self._parse_int(item_data.get("runtime") or item_data.get("Runtime"))
        if "rating" in item_data or "Rating" in item_data:
            existing_item.rating = self._parse_float(item_data.get("rating") or item_data.get("Rating"))

        existing_item.updated_at = datetime.utcnow()
        session.add(existing_item)

    def _delete_existing_item(self, session: Any, item: MediaItem) -> None:
        """Delete an existing media item and its related data."""
        session.delete(item)

    def _detect_conflicts(self, data: list[dict[str, Any]]) -> list[ImportConflict]:
        """Detect potential conflicts in import data."""
        conflicts: list[ImportConflict] = []

        with transactional_context() as uow:
            session = uow.session

            for idx, item_data in enumerate(data, start=1):
                title = item_data.get("title") or item_data.get("Title")
                if not title:
                    conflicts.append(
                        ImportConflict(
                            row_index=idx,
                            title="",
                            conflict_type="missing_title",
                            details="Title field is required",
                        )
                    )
                    continue

                # Check for duplicate by ID
                item_id = item_data.get("id") or item_data.get("ID")
                if item_id:
                    stmt = select(MediaItem).where(MediaItem.id == int(item_id))
                    existing = session.exec(stmt).first()
                    if existing:
                        conflicts.append(
                            ImportConflict(
                                row_index=idx,
                                title=title,
                                conflict_type="duplicate_id",
                                existing_item_id=existing.id,
                                details=f"Item with ID {item_id} already exists",
                            )
                        )

                # Check for duplicate by title/year
                year = item_data.get("year") or item_data.get("Year")
                stmt = select(MediaItem).where(MediaItem.title == title)
                if year:
                    stmt = stmt.where(MediaItem.year == year)
                existing = session.exec(stmt).first()
                if existing:
                    conflicts.append(
                        ImportConflict(
                            row_index=idx,
                            title=title,
                            conflict_type="duplicate_title",
                            existing_item_id=existing.id,
                            details=f"Item with title '{title}' ({year}) already exists",
                        )
                    )

        return conflicts

    def _parse_int(self, value: Any) -> int | None:
        """Parse an integer value safely."""
        if value is None or value == "":
            return None
        try:
            return int(value)
        except (ValueError, TypeError):
            return None

    def _parse_float(self, value: Any) -> float | None:
        """Parse a float value safely."""
        if value is None or value == "":
            return None
        try:
            return float(value)
        except (ValueError, TypeError):
            return None
